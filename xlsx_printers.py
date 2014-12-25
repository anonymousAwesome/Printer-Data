#this script requires openpyxl.

from openpyxl import load_workbook

wb=load_workbook("PrinterData.xlsx", use_iterators=True)
ws=wb.active

desired_lowercase_results={
	'IsComplete': "y",
	'IsValidItem': "y",
	'StaplesPrinterUse': ["medium_office", "large_office", "high_productivity"],
	'SupportsScanning': "y",
#	'SupportsCopying': "n",
	'SupportedPageSize': "tabloid",
	'SupportsWiFiPrinting': "y",
	'HasTouchScreenInterface': "y"
}



def compare_dicts(xlsx_dict, desired_dict):
	'''Compares the dictionary of desired requirements with a dictionary
	of one item gleaned from the excel file.  If all of the requirements
	are met, the function returns True.  Otherwise, it returns False.
	
	For requirements that have multiple options (e.g. medium_office or 
	large_office), only one of those criteria needs to be met in order
	for the test to return True.'''
	
	list=[]
	for key in desired_lowercase_results:
		if not xlsx_dict[key]:
			return False
		else:
			list.append(xlsx_dict[key].lower() in desired_lowercase_results[key])
	if False in list:
		return False
	else:
		return True

number_of_dictionaries=0

for row in ws.iter_rows():

	dict={
		'SKU': row[0].value,
		'IsComplete': row[2].value,
		'IsValidItem': row[3].value,
		'ModelNumber': row[5].value,
		'ProductTitle': row[7].value,
		'StaplesPrinterUse': row[10].value,
		'SupportsScanning': row[17].value,
		'SupportsCopying': row[18].value,
		'SupportedPageSize': row[19].value,
		'SupportsWiFiPrinting': row[21].value,
		'HasTouchScreenInterface': row[22].value
	}

	
	if compare_dicts(dict,desired_lowercase_results):
		for key in dict:
			print key, ": ", dict[key]
		print ""
		number_of_dictionaries+=1

print number_of_dictionaries, "results from the excel spreadsheet."
